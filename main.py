import telebot
from selenium import webdriver
import mysql.connector
from bs4 import BeautifulSoup
import os
import openpyxl


token = '1792508263:AAGbFGJcYb8VedM8Sl04HsizlZu4vd03Sa8'

bot = telebot.TeleBot(token)

my_db = mysql.connector.connect(
  host="localhost",
  user="root",
  passwd="",
  database="telebot_users"
)

chromedriver = 'C:/Chromedriver/chromedriver.exe'
options = webdriver.ChromeOptions()
options.add_argument('headless')  # для открытия headless-браузера
browser = webdriver.Chrome(executable_path=chromedriver, options=options)


def private_users(user_id):
    global my_db
    cursor = my_db.cursor()
    sql = 'select user_id from private_users where user_id = ' + str(user_id)
    cursor.execute(sql)
    result = cursor.fetchone()
    return result


def parse_phones(user_id, url, file_name):
    wb = openpyxl.Workbook()
    wb.create_sheet(title='Sheet1', index=0)
    sheet = wb['Sheet1']
    i = 1
    j = 1
    cell = sheet.cell(row=j, column=1)
    cell.value = 'ID'
    cell = sheet.cell(row=j, column=2)
    cell.value = 'Product_ID'
    cell = sheet.cell(row=j, column=3)
    cell.value = 'NAME'
    cell = sheet.cell(row=j, column=4)
    cell.value = 'PRICE'
    cell = sheet.cell(row=j, column=5)
    cell.value = 'LINK'

    while i < 19:
        url = str(url).format(i)

        browser.get(url)
        html = browser.page_source
        result = []

        soup = BeautifulSoup(html, 'lxml')
        divs = soup.find('div', class_="item-cards-grid__cards").find_all('div', class_="item-card ddl_product ddl_product_link undefined")

        for d in divs:
            price = d.find('span', class_="item-card__prices-price").text.replace(' ', '').replace('₸', '')
            name = d.find('a', class_="item-card__name-link").text
            product_id = d.get('data-product-id')
            link = d.find('a', class_='item-card__name-link').get('href')

            j += 1
            value_price = price
            value_product_id = product_id
            value_name = name
            value_link = link
            cell = sheet.cell(row=j, column=1)
            cell.value = j
            cell = sheet.cell(row=j, column=2)
            cell.value = value_product_id
            cell = sheet.cell(row=j, column=3)
            cell.value = value_name
            cell = sheet.cell(row=j, column=4)
            cell.value = value_price
            cell = sheet.cell(row=j, column=5)
            cell.value = value_link
        i += 1
        wb.save('C:/Users/Bruce/PycharmProjects/untitled/pythonProject/Telebot_parcer/' + str(user_id) +
                str(file_name) + str(user_id) + '.xlsx')


@bot.message_handler(content_types=["text"])
def handle_text(message):
    global menu

    if message.text.lower() == "/start":
        menu = "start"
        text = "Это парсер-бот магазина kaspi.kz\n"
        text = text + "1 - Если вы являетесь нашим клиентом\n"
        text = text + "2 - Если вы хотите зарегистрироваться\n"
        bot.send_message(message.chat.id, text)
        menu = "authorization"

    else:
        if menu == "authorization":
            if message.text.lower() == "1":
                client = private_users(message.from_user.id)
                if client is not None:
                    text = "Выберите сегмент для парсинга\n"
                    text = text + "1 - Телефоны Apple\n"
                    text = text + "2 - Ноутбуки\n"
                    msg = bot.send_message(message.chat.id, text)
                    bot.register_next_step_handler(msg, menu_parse)

                else:
                    text = "Вас не нашли в нашей базе данных\n"
                    text = text + "Напишите https://t.me/G33p0x для регистрации\n"
                    text = text + "Нажмите /start для продолжения"
                    bot.send_message(message.chat.id, text)
                    menu = "start"

            elif message.text.lower() == "2":
                text = "Напишите https://t.me/G33p0x для регистрации\n"
                text = text + "Нажмите /start для продолжения"
                bot.send_message(message.chat.id, text)

            else:
                text = "Вы ввели неверно\n"
                text = text + "Нажмите /start для продолжения"
                bot.send_message(message.chat.id, text)


def menu_parse(message):
    menu = "main"
    if menu == "main":
        if message.text.lower() == "1":
            bot.reply_to(message, 'Подождите пару минут пока бот парсит данные')
            path = "C:/Users/Bruce/PycharmProjects/untitled/pythonProject/Telebot_parcer/" + str(
                message.from_user.id)
            os.mkdir(path)
            url = "https://kaspi.kz/shop/c/smartphones/?q=%3AproductClass%3AApple+iPhone&page={}"
            file_name = "/apple_data_"
            parse_phones(message.from_user.id, url, file_name)
            bot.send_message(message.chat.id, "Ваш файл готов")
            file = open(path + file_name + str(message.from_user.id) + '.xlsx', 'rb')
            bot.send_document(message.chat.id, file)
            file.close()
            os.remove(path + file_name + str(message.from_user.id) + ".xlsx")
            os.rmdir(path)
            bot.send_message(message.chat.id, "Нажмите /start для продолжения")

        elif message.text.lower() == "2":
            bot.reply_to(message, 'Подождите пару минут пока бот парсит данные')
            path = "C:/Users/Bruce/PycharmProjects/untitled/pythonProject/Telebot_parcer/" + str(
                message.from_user.id)
            os.mkdir(path)
            url = "https://kaspi.kz/shop/c/notebooks/?q=%3AproductClass%3A%D0%98%D0%B3%D1%80%D0%BE%D0%B2%D1%8B%D0%B5&amp;page={}"
            file_name = "/laptop_data_"
            parse_phones(message.from_user.id, url, file_name)
            bot.send_message(message.chat.id, "Ваш файл готов")
            file = open(path + file_name + str(message.from_user.id) + '.xlsx', 'rb')
            bot.send_document(message.chat.id, file)
            file.close()
            os.remove(path + file_name + str(message.from_user.id) + ".xlsx")
            os.rmdir(path)
            bot.send_message(message.chat.id, "Нажмите /start для продолжения")
        else:
            text = "Вы ввели неверно\n"
            text = text + "Нажмите /start для продолжения"
            bot.send_message(message.chat.id, text)


bot.polling(none_stop=True, interval=0)


